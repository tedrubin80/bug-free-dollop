#!/usr/bin/env python3
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

def create_interactive_gear_calculator():
    """Create an interactive gear calculator with item selection from database"""

    # Read the scraped items data
    try:
        df = pd.read_csv('fallensword_items.csv')
    except FileNotFoundError:
        print("Error: fallensword_items.csv not found. Please run the scraper first.")
        return

    # Clean numeric columns
    numeric_cols = ['Level', 'Attack', 'Defense', 'Armor', 'Damage', 'HP', 'Stamina']
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)

    # Create workbook
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    total_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    total_font = Font(color="FFFFFF", bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))

    # Sheet 1: Instructions
    ws_instructions = wb.create_sheet("Instructions")
    instructions = [
        ["FALLEN SWORD GEAR CALCULATOR - INTERACTIVE VERSION"],
        [""],
        ["HOW TO USE:"],
        ["1. Go to 'Gear Calculator' sheet"],
        ["2. Click on any cell in the 'Selected Item Name' column (Column B)"],
        ["3. Start typing an item name - suggestions will appear as you type"],
        ["4. Select the item you want from the dropdown"],
        ["5. The stats will automatically populate"],
        ["6. Your total stats are shown at the bottom"],
        [""],
        ["FEATURES:"],
        ["• Auto-complete item names from the database"],
        ["• Instant stat calculation when you select items"],
        ["• Color coding for different stat types"],
        ["• Item comparison tool to compare two items"],
        ["• Best items filtered by your level"],
        [""],
        ["SHEETS INCLUDED:"],
        ["• Items Database - All 3,963 items from the game"],
        ["• Gear Calculator - Build your equipment loadout"],
        ["• Item Comparison - Compare two items side by side"],
        ["• Items by Type - Browse items filtered by equipment type"],
        ["• My Level Items - Items filtered to your character level"],
        [""],
        ["IMPORTING TO GOOGLE SHEETS:"],
        ["1. Open Google Sheets (sheets.google.com)"],
        ["2. Create a new spreadsheet"],
        ["3. File → Import → Upload this file"],
        ["4. Choose 'Replace spreadsheet'"],
        ["5. After import, set up Data Validation:"],
        ["   - Select cells B2:B11 in Gear Calculator"],
        ["   - Data → Data validation"],
        ["   - Criteria: List from range → 'Items Database'!A:A"],
        ["   - This enables the dropdown with all item names"]
    ]

    for row in instructions:
        ws_instructions.append(row)

    # Format instructions
    ws_instructions.column_dimensions['A'].width = 80
    ws_instructions['A1'].font = Font(size=16, bold=True, color="4472C4")

    # Sheet 2: Items Database
    ws_items = wb.create_sheet("Items Database")

    # Add headers
    headers = ['Name', 'Level', 'Type', 'Rarity', 'Attack', 'Defense', 'Armor', 'Damage', 'HP', 'Stamina']
    ws_items.append(headers)

    # Add data
    for _, row in df.iterrows():
        ws_items.append([
            row.get('Name', ''),
            row.get('Level', 0),
            row.get('Type', ''),
            row.get('Rarity', ''),
            row.get('Attack', 0),
            row.get('Defense', 0),
            row.get('Armor', 0),
            row.get('Damage', 0),
            row.get('HP', 0),
            row.get('Stamina', 0)
        ])

    # Format headers
    for cell in ws_items[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Auto-adjust column widths
    for column in ws_items.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws_items.column_dimensions[column_letter].width = adjusted_width

    # Sheet 3: Gear Calculator
    ws_calc = wb.create_sheet("Gear Calculator")

    # Add title
    ws_calc.merge_cells('A1:J1')
    ws_calc['A1'] = 'GEAR CALCULATOR - SELECT ITEMS FROM DATABASE'
    ws_calc['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws_calc['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws_calc['A1'].alignment = Alignment(horizontal='center')

    # Add headers
    calc_headers = ['Equipment Slot', 'Selected Item Name', 'Type', 'Level', 'Attack', 'Defense', 'Armor', 'Damage', 'HP', 'Stamina']
    ws_calc.append([''])  # Empty row
    ws_calc.append(calc_headers)

    # Format headers
    for cell in ws_calc[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Add equipment slots with formulas
    equipment_slots = [
        ('Helmet', 'Helmet'),
        ('Armor', 'Armor'),
        ('Gloves', 'Gloves'),
        ('Boots', 'Boots'),
        ('Weapon', 'Weapon'),
        ('Shield', 'Shield'),
        ('Ring 1', 'Ring'),
        ('Ring 2', 'Ring'),
        ('Amulet', 'Amulet'),
        ('Rune', 'Rune')
    ]

    start_row = 4
    for i, (slot, type_filter) in enumerate(equipment_slots):
        row_num = start_row + i

        # Equipment slot name
        ws_calc.cell(row=row_num, column=1, value=slot)
        ws_calc.cell(row=row_num, column=1).font = Font(bold=True)
        ws_calc.cell(row=row_num, column=1).border = border

        # Selected item name (user input)
        ws_calc.cell(row=row_num, column=2, value='')
        ws_calc.cell(row=row_num, column=2).fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
        ws_calc.cell(row=row_num, column=2).border = border

        # Type (VLOOKUP)
        ws_calc.cell(row=row_num, column=3, value=f'=IFERROR(VLOOKUP(B{row_num},\'Items Database\'!A:J,3,FALSE),"")')
        ws_calc.cell(row=row_num, column=3).border = border

        # Level (VLOOKUP)
        ws_calc.cell(row=row_num, column=4, value=f'=IFERROR(VLOOKUP(B{row_num},\'Items Database\'!A:J,2,FALSE),0)')
        ws_calc.cell(row=row_num, column=4).border = border

        # Attack (VLOOKUP)
        ws_calc.cell(row=row_num, column=5, value=f'=IFERROR(VLOOKUP(B{row_num},\'Items Database\'!A:J,5,FALSE),0)')
        ws_calc.cell(row=row_num, column=5).border = border

        # Defense (VLOOKUP)
        ws_calc.cell(row=row_num, column=6, value=f'=IFERROR(VLOOKUP(B{row_num},\'Items Database\'!A:J,6,FALSE),0)')
        ws_calc.cell(row=row_num, column=6).border = border

        # Armor (VLOOKUP)
        ws_calc.cell(row=row_num, column=7, value=f'=IFERROR(VLOOKUP(B{row_num},\'Items Database\'!A:J,7,FALSE),0)')
        ws_calc.cell(row=row_num, column=7).border = border

        # Damage (VLOOKUP)
        ws_calc.cell(row=row_num, column=8, value=f'=IFERROR(VLOOKUP(B{row_num},\'Items Database\'!A:J,8,FALSE),0)')
        ws_calc.cell(row=row_num, column=8).border = border

        # HP (VLOOKUP)
        ws_calc.cell(row=row_num, column=9, value=f'=IFERROR(VLOOKUP(B{row_num},\'Items Database\'!A:J,9,FALSE),0)')
        ws_calc.cell(row=row_num, column=9).border = border

        # Stamina (VLOOKUP)
        ws_calc.cell(row=row_num, column=10, value=f'=IFERROR(VLOOKUP(B{row_num},\'Items Database\'!A:J,10,FALSE),0)')
        ws_calc.cell(row=row_num, column=10).border = border

    # Add total row
    total_row = start_row + len(equipment_slots) + 1
    ws_calc.cell(row=total_row, column=1, value='TOTAL STATS')
    ws_calc.cell(row=total_row, column=1).fill = total_fill
    ws_calc.cell(row=total_row, column=1).font = total_font
    ws_calc.cell(row=total_row, column=1).border = border

    # Merge cells for total label
    ws_calc.merge_cells(f'B{total_row}:C{total_row}')
    ws_calc.cell(row=total_row, column=2, value='Your Combined Stats')
    ws_calc.cell(row=total_row, column=2).fill = total_fill
    ws_calc.cell(row=total_row, column=2).font = total_font
    ws_calc.cell(row=total_row, column=2).alignment = Alignment(horizontal='center')
    ws_calc.cell(row=total_row, column=2).border = border

    # Total formulas
    ws_calc.cell(row=total_row, column=4, value=f'=MAX(D{start_row}:D{start_row+9})')
    ws_calc.cell(row=total_row, column=5, value=f'=SUM(E{start_row}:E{start_row+9})')
    ws_calc.cell(row=total_row, column=6, value=f'=SUM(F{start_row}:F{start_row+9})')
    ws_calc.cell(row=total_row, column=7, value=f'=SUM(G{start_row}:G{start_row+9})')
    ws_calc.cell(row=total_row, column=8, value=f'=SUM(H{start_row}:H{start_row+9})')
    ws_calc.cell(row=total_row, column=9, value=f'=SUM(I{start_row}:I{start_row+9})')
    ws_calc.cell(row=total_row, column=10, value=f'=SUM(J{start_row}:J{start_row+9})')

    for col in range(4, 11):
        ws_calc.cell(row=total_row, column=col).fill = total_fill
        ws_calc.cell(row=total_row, column=col).font = total_font
        ws_calc.cell(row=total_row, column=col).border = border

    # Adjust column widths
    ws_calc.column_dimensions['A'].width = 15
    ws_calc.column_dimensions['B'].width = 30
    for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        ws_calc.column_dimensions[col].width = 10

    # Add note about data validation
    ws_calc.cell(row=total_row+2, column=1, value="NOTE: After importing to Google Sheets, set up Data Validation on column B (yellow cells)")
    ws_calc.cell(row=total_row+3, column=1, value="Select B4:B13 → Data → Data validation → List from 'Items Database'!A:A")
    ws_calc.cell(row=total_row+2, column=1).font = Font(italic=True, color="FF0000")
    ws_calc.cell(row=total_row+3, column=1).font = Font(italic=True, color="FF0000")

    # Sheet 4: Item Comparison
    ws_compare = wb.create_sheet("Item Comparison")

    # Title
    ws_compare.merge_cells('A1:K1')
    ws_compare['A1'] = 'ITEM COMPARISON TOOL'
    ws_compare['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws_compare['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws_compare['A1'].alignment = Alignment(horizontal='center')

    # Headers
    ws_compare.append([''])
    compare_headers = ['', 'Item Name', 'Type', 'Level', 'Attack', 'Defense', 'Armor', 'Damage', 'HP', 'Stamina', 'Total Stats']
    ws_compare.append(compare_headers)

    for cell in ws_compare[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Item 1
    ws_compare.cell(row=4, column=1, value='Item 1')
    ws_compare.cell(row=4, column=1).font = Font(bold=True)
    ws_compare.cell(row=4, column=2, value='')  # User input
    ws_compare.cell(row=4, column=2).fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

    for col in range(3, 11):
        formula_col = col - 1 if col > 2 else col
        ws_compare.cell(row=4, column=col, value=f'=IFERROR(VLOOKUP(B4,\'Items Database\'!A:J,{formula_col},FALSE),0)')

    ws_compare.cell(row=4, column=11, value='=SUM(E4:J4)')

    # Item 2
    ws_compare.cell(row=5, column=1, value='Item 2')
    ws_compare.cell(row=5, column=1).font = Font(bold=True)
    ws_compare.cell(row=5, column=2, value='')  # User input
    ws_compare.cell(row=5, column=2).fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

    for col in range(3, 11):
        formula_col = col - 1 if col > 2 else col
        ws_compare.cell(row=5, column=col, value=f'=IFERROR(VLOOKUP(B5,\'Items Database\'!A:J,{formula_col},FALSE),0)')

    ws_compare.cell(row=5, column=11, value='=SUM(E5:J5)')

    # Difference row
    ws_compare.cell(row=6, column=1, value='Difference')
    ws_compare.cell(row=6, column=1).font = Font(bold=True, color="FF0000")
    ws_compare.cell(row=6, column=2, value='(Item 2 - Item 1)')

    for col in range(3, 12):
        if col == 3:  # Type column
            ws_compare.cell(row=6, column=col, value='-')
        else:
            ws_compare.cell(row=6, column=col, value=f'={get_column_letter(col)}5-{get_column_letter(col)}4')
            ws_compare.cell(row=6, column=col).font = Font(bold=True)

    # Adjust column widths
    ws_compare.column_dimensions['A'].width = 12
    ws_compare.column_dimensions['B'].width = 30
    for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
        ws_compare.column_dimensions[col].width = 10

    # Sheet 5: Items by Type (filtered views)
    ws_types = wb.create_sheet("Items by Type")

    # Get unique types
    item_types = df['Type'].dropna().unique()

    # Create sections for each type (showing top 20 items of each type)
    current_row = 1
    for item_type in sorted(item_types)[:10]:  # Limit to first 10 types for space
        # Type header
        ws_types.cell(row=current_row, column=1, value=f'{item_type.upper()} ITEMS')
        ws_types.cell(row=current_row, column=1).font = Font(size=12, bold=True, color="FFFFFF")
        ws_types.cell(row=current_row, column=1).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws_types.merge_cells(f'A{current_row}:J{current_row}')

        current_row += 1

        # Headers
        type_headers = ['Name', 'Level', 'Rarity', 'Attack', 'Defense', 'Armor', 'Damage', 'HP', 'Stamina', 'Total']
        for col, header in enumerate(type_headers, 1):
            ws_types.cell(row=current_row, column=col, value=header)
            ws_types.cell(row=current_row, column=col).font = Font(bold=True)
            ws_types.cell(row=current_row, column=col).fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

        current_row += 1

        # Get top items of this type
        type_items = df[df['Type'] == item_type].copy()
        type_items['Total'] = type_items[['Attack', 'Defense', 'Armor', 'Damage', 'HP', 'Stamina']].sum(axis=1)
        type_items = type_items.nlargest(20, 'Total')

        # Add items
        for _, item in type_items.iterrows():
            ws_types.cell(row=current_row, column=1, value=item['Name'])
            ws_types.cell(row=current_row, column=2, value=item['Level'])
            ws_types.cell(row=current_row, column=3, value=item['Rarity'])
            ws_types.cell(row=current_row, column=4, value=item['Attack'])
            ws_types.cell(row=current_row, column=5, value=item['Defense'])
            ws_types.cell(row=current_row, column=6, value=item['Armor'])
            ws_types.cell(row=current_row, column=7, value=item['Damage'])
            ws_types.cell(row=current_row, column=8, value=item['HP'])
            ws_types.cell(row=current_row, column=9, value=item['Stamina'])
            ws_types.cell(row=current_row, column=10, value=item['Total'])
            current_row += 1

        current_row += 2  # Space between types

    # Save the workbook
    wb.save('fallensword_interactive_calculator.xlsx')
    print("Interactive gear calculator created successfully!")

    # Create a simpler version for Google Sheets with clear instructions
    create_google_sheets_version(df)

def create_google_sheets_version(df):
    """Create a simplified version optimized for Google Sheets"""

    with pd.ExcelWriter('fallensword_google_sheets_calculator.xlsx', engine='openpyxl') as writer:

        # Instructions sheet
        instructions_df = pd.DataFrame({
            'SETUP INSTRUCTIONS FOR GOOGLE SHEETS': [
                '1. IMPORT THIS FILE:',
                '   - Open Google Sheets',
                '   - File → Import → Upload this file',
                '   - Choose "Replace spreadsheet"',
                '',
                '2. SET UP ITEM SELECTION DROPDOWN:',
                '   - Go to "Gear Calculator" sheet',
                '   - Select cells B4:B13 (the yellow cells)',
                '   - Click Data → Data validation',
                '   - Add rule:',
                '     • Criteria: "Dropdown (from a range)"',
                '     • Range: Items Database!A:A',
                '     • Click "Done"',
                '',
                '3. HOW TO USE:',
                '   - Click any yellow cell in Gear Calculator',
                '   - Type to search or click the dropdown arrow',
                '   - Select your item',
                '   - Stats auto-calculate!',
                '',
                '4. FOR ITEM COMPARISON:',
                '   - Go to "Item Comparison" sheet',
                '   - Set up data validation on B4:B5 same as above',
                '   - Select two items to compare',
                '',
                'TIPS:',
                '• You can type partial item names to filter',
                '• Use Ctrl+F to search in Items Database',
                '• Create multiple copies for different builds'
            ]
        })
        instructions_df.to_excel(writer, sheet_name='SETUP_INSTRUCTIONS', index=False)

        # Items Database
        df.to_excel(writer, sheet_name='Items Database', index=False)

        # Gear Calculator with formulas
        calc_data = []
        for i, slot in enumerate(['Helmet', 'Armor', 'Gloves', 'Boots', 'Weapon',
                                  'Shield', 'Ring 1', 'Ring 2', 'Amulet', 'Rune']):
            row = i + 4
            calc_data.append({
                'Equipment Slot': slot,
                'Selected Item Name': '',  # Will be dropdown
                'Type': f'=IFERROR(VLOOKUP(B{row},\'Items Database\'!A:K,3,FALSE),"")',
                'Level': f'=IFERROR(VLOOKUP(B{row},\'Items Database\'!A:K,2,FALSE),0)',
                'Attack': f'=IFERROR(VLOOKUP(B{row},\'Items Database\'!A:K,5,FALSE),0)',
                'Defense': f'=IFERROR(VLOOKUP(B{row},\'Items Database\'!A:K,6,FALSE),0)',
                'Armor': f'=IFERROR(VLOOKUP(B{row},\'Items Database\'!A:K,7,FALSE),0)',
                'Damage': f'=IFERROR(VLOOKUP(B{row},\'Items Database\'!A:K,8,FALSE),0)',
                'HP': f'=IFERROR(VLOOKUP(B{row},\'Items Database\'!A:K,9,FALSE),0)',
                'Stamina': f'=IFERROR(VLOOKUP(B{row},\'Items Database\'!A:K,10,FALSE),0)'
            })

        # Add total row
        calc_data.append({
            'Equipment Slot': 'TOTAL STATS',
            'Selected Item Name': '',
            'Type': '',
            'Level': '=MAX(D4:D13)',
            'Attack': '=SUM(E4:E13)',
            'Defense': '=SUM(F4:F13)',
            'Armor': '=SUM(G4:G13)',
            'Damage': '=SUM(H4:H13)',
            'HP': '=SUM(I4:I13)',
            'Stamina': '=SUM(J4:J13)'
        })

        calc_df = pd.DataFrame(calc_data)
        calc_df.to_excel(writer, sheet_name='Gear Calculator', index=False, startrow=2)

        # Write title
        worksheet = writer.sheets['Gear Calculator']
        worksheet.cell(row=1, column=1).value = 'GEAR CALCULATOR - Select items in column B (set up dropdown as per instructions)'

        # Item Comparison
        comparison_data = [
            {
                'Item': 'Item 1',
                'Name': '',
                'Type': '=IFERROR(VLOOKUP(B4,\'Items Database\'!A:K,3,FALSE),"")',
                'Level': '=IFERROR(VLOOKUP(B4,\'Items Database\'!A:K,2,FALSE),0)',
                'Attack': '=IFERROR(VLOOKUP(B4,\'Items Database\'!A:K,5,FALSE),0)',
                'Defense': '=IFERROR(VLOOKUP(B4,\'Items Database\'!A:K,6,FALSE),0)',
                'Armor': '=IFERROR(VLOOKUP(B4,\'Items Database\'!A:K,7,FALSE),0)',
                'Damage': '=IFERROR(VLOOKUP(B4,\'Items Database\'!A:K,8,FALSE),0)',
                'HP': '=IFERROR(VLOOKUP(B4,\'Items Database\'!A:K,9,FALSE),0)',
                'Stamina': '=IFERROR(VLOOKUP(B4,\'Items Database\'!A:K,10,FALSE),0)',
                'Total': '=SUM(E4:J4)'
            },
            {
                'Item': 'Item 2',
                'Name': '',
                'Type': '=IFERROR(VLOOKUP(B5,\'Items Database\'!A:K,3,FALSE),"")',
                'Level': '=IFERROR(VLOOKUP(B5,\'Items Database\'!A:K,2,FALSE),0)',
                'Attack': '=IFERROR(VLOOKUP(B5,\'Items Database\'!A:K,5,FALSE),0)',
                'Defense': '=IFERROR(VLOOKUP(B5,\'Items Database\'!A:K,6,FALSE),0)',
                'Armor': '=IFERROR(VLOOKUP(B5,\'Items Database\'!A:K,7,FALSE),0)',
                'Damage': '=IFERROR(VLOOKUP(B5,\'Items Database\'!A:K,8,FALSE),0)',
                'HP': '=IFERROR(VLOOKUP(B5,\'Items Database\'!A:K,9,FALSE),0)',
                'Stamina': '=IFERROR(VLOOKUP(B5,\'Items Database\'!A:K,10,FALSE),0)',
                'Total': '=SUM(E5:J5)'
            },
            {
                'Item': 'Difference',
                'Name': '(Item 2 - Item 1)',
                'Type': '-',
                'Level': '=D5-D4',
                'Attack': '=E5-E4',
                'Defense': '=F5-F4',
                'Armor': '=G5-G4',
                'Damage': '=H5-H4',
                'HP': '=I5-I4',
                'Stamina': '=J5-J4',
                'Total': '=K5-K4'
            }
        ]

        comparison_df = pd.DataFrame(comparison_data)
        comparison_df.to_excel(writer, sheet_name='Item Comparison', index=False, startrow=2)

        # Write title
        worksheet = writer.sheets['Item Comparison']
        worksheet.cell(row=1, column=1).value = 'ITEM COMPARISON - Compare two items side by side'

def main():
    print("\n🎮 Creating Interactive Fallen Sword Gear Calculator...")
    print("=" * 60)

    create_interactive_gear_calculator()

    print("\n✅ Files created successfully!")
    print("\n📁 Created files:")
    print("  • fallensword_interactive_calculator.xlsx")
    print("  • fallensword_google_sheets_calculator.xlsx")

    print("\n📊 TO USE IN GOOGLE SHEETS:")
    print("=" * 40)
    print("1. Open Google Sheets (sheets.google.com)")
    print("2. File → Import → Upload 'fallensword_google_sheets_calculator.xlsx'")
    print("3. Follow the SETUP_INSTRUCTIONS sheet")
    print("\nKey step: Set up Data Validation dropdowns:")
    print("  - Select the yellow cells (B4:B13) in Gear Calculator")
    print("  - Data → Data validation → List from 'Items Database'!A:A")
    print("\n✨ Then just click any equipment slot and select your item!")
    print("   The stats will auto-calculate from the database!")

if __name__ == "__main__":
    main()